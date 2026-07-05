import sqlite3
from openpyxl import Workbook
import matplotlib.pyplot as plt

# ---------------- DATABASE ----------------

conn = sqlite3.connect("erp.db")
cursor = conn.cursor()

# ==========================================
# CREATE TABLES
# ==========================================

# Employee Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS employees(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    department TEXT,
    salary REAL
)
""")

# Customer Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS customers(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    phone TEXT
)
""")

# Product Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS products(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    price REAL,
    quantity INTEGER
)
""")

# Inventory Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS inventory(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    product_name TEXT,
    quantity INTEGER,
    location TEXT
)
""")

# Supplier Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS suppliers(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    phone TEXT,
    email TEXT,
    company TEXT
)
""")

# Payments Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS payments(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer TEXT,
    amount REAL,
    payment_method TEXT,
    payment_date TEXT
)
""")

# Sales Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS sales(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    customer TEXT,
    product TEXT,
    quantity INTEGER,
    price REAL,
    total REAL
)
""")

# Attendance Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS attendance(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_name TEXT,
    date TEXT,
    status TEXT
)
""")

# PAYROLL TABLE
cursor.execute("""
CREATE TABLE IF NOT EXISTS payroll(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_name TEXT,
    basic_salary REAL,
    bonus REAL,
    deduction REAL,
    net_salary REAL
)
""")

conn.commit()

# ---------------- LOGIN ----------------
USERNAME = "admin"
PASSWORD = "1234"

print("=" * 40)
print("        ERP MANAGEMENT SYSTEM")
print("=" * 40)

username = input("Username : ")
password = input("Password : ")

if username != USERNAME or password != PASSWORD:
    print("Invalid Login")
    exit()

print("\nLogin Successful")

# ---------------- EMPLOYEE ----------------

def add_employee():

    name = input("Employee Name : ")
    department = input("Department : ")
    salary = float(input("Salary : "))

    cursor.execute(
        "INSERT INTO employees(name,department,salary) VALUES(?,?,?)",
        (name, department, salary)
    )

    conn.commit()

    print("Employee Added Successfully")


def view_employee():

    cursor.execute("SELECT * FROM employees")

    rows = cursor.fetchall()

    print("\nEMPLOYEE LIST")

    print("-" * 50)

    for row in rows:
        print("ID :", row[0])
        print("Name :", row[1])
        print("Department :", row[2])
        print("Salary :", row[3])
        print("-" * 50)


def update_employee():

    emp_id = int(input("Employee ID : "))

    name = input("New Name : ")
    department = input("New Department : ")
    salary = float(input("New Salary : "))

    cursor.execute(
        """
        UPDATE employees
        SET name=?, department=?, salary=?
        WHERE id=?
        """,
        (name, department, salary, emp_id)
    )

    conn.commit()

    print("Employee Updated")


def delete_employee():

    emp_id = int(input("Employee ID : "))

    cursor.execute(
        "DELETE FROM employees WHERE id=?",
        (emp_id,)
    )

    conn.commit()

    print("Employee Deleted")


# ---------------- CUSTOMER ----------------

def add_customer():

    name = input("Customer Name : ")
    phone = input("Phone : ")

    cursor.execute(
        "INSERT INTO customers(name,phone) VALUES(?,?)",
        (name, phone)
    )

    conn.commit()

    print("Customer Added")


def view_customer():

    cursor.execute("SELECT * FROM customers")

    rows = cursor.fetchall()

    print("\nCUSTOMER LIST")

    for row in rows:

        print("ID :", row[0])
        print("Name :", row[1])
        print("Phone :", row[2])
        print("-" * 40)


# ---------------- SUPPLIER ----------------

def add_supplier():

    name = input("Supplier Name : ")
    phone = input("Phone : ")
    email = input("Email : ")
    company = input("Company : ")

    cursor.execute(
        "INSERT INTO suppliers(name,phone,email,company) VALUES(?,?,?,?)",
        (name, phone, email, company)
    )

    conn.commit()

    print("Supplier Added Successfully")


def view_supplier():

    cursor.execute("SELECT * FROM suppliers")

    rows = cursor.fetchall()

    print("\nSUPPLIER LIST")

    print("-" * 50)

    for row in rows:

        print("ID :", row[0])
        print("Name :", row[1])
        print("Phone :", row[2])
        print("Email :", row[3])
        print("Company :", row[4])
        print("-" * 50)


def update_supplier():

    supp_id = int(input("Supplier ID : "))

    name = input("New Name : ")
    phone = input("New Phone : ")
    email = input("New Email : ")
    company = input("New Company : ")

    cursor.execute(
        """
        UPDATE suppliers
        SET name=?, phone=?, email=?, company=?
        WHERE id=?
        """,
        (name, phone, email, company, supp_id)
    )

    conn.commit()

    print("Supplier Updated Successfully")


def delete_supplier():

    supp_id = int(input("Supplier ID : "))

    cursor.execute(
        "DELETE FROM suppliers WHERE id=?",
        (supp_id,)
    )

    conn.commit()

    print("Supplier Deleted Successfully")

def search_supplier():

    name = input("Supplier Name : ")

    cursor.execute(
        "SELECT * FROM suppliers WHERE name LIKE ?",
        ("%" + name + "%",)
    )

    rows = cursor.fetchall()

    if len(rows)==0:
        print("Supplier Not Found")
        return

    for row in rows:
        print(row)


# ---------------- PRODUCT ----------------

def add_product():

    name = input("Product Name : ")
    price = float(input("Price : "))
    quantity = int(input("Quantity : "))

    cursor.execute(
        """
        INSERT INTO products(name,price,quantity)
        VALUES(?,?,?)
        """,
        (name, price, quantity)
    )

    conn.commit()

    print("Product Added")


def view_product():

    cursor.execute("SELECT * FROM products")

    rows = cursor.fetchall()

    print("\nPRODUCT LIST")

    for row in rows:

        print("ID :", row[0])
        print("Name :", row[1])
        print("Price :", row[2])
        print("Quantity :", row[3])
        print("-" * 40)


def search_product():

    name = input("Product Name : ")

    cursor.execute(
        "SELECT * FROM products WHERE name LIKE ?",
        ("%" + name + "%",)
    )

    rows = cursor.fetchall()

    if len(rows) == 0:
        print("No Product Found")
        return

    print("\n======= SEARCH RESULT =======")

    for row in rows:

        print("ID :", row[0])
        print("Name :", row[1])
        print("Price :", row[2])
        print("Quantity :", row[3])
        print("-" * 40)


def update_product():

    prod_id = int(input("Product ID : "))

    name = input("New Name : ")
    price = float(input("New Price : "))
    quantity = int(input("New Quantity : "))

    cursor.execute("""
    UPDATE products
    SET name=?, price=?, quantity=?
    WHERE id=?
    """, (name, price, quantity, prod_id))

    conn.commit()

    print("Product Updated Successfully")


def delete_product():

    prod_id = int(input("Product ID : "))

    cursor.execute(
        "DELETE FROM products WHERE id=?",
        (prod_id,)
    )

    conn.commit()

    print("Product Deleted Successfully")


# ======================================
# ADD INVENTORY FUNCTIONS HERE
# ======================================

def add_inventory():

    product = input("Product Name : ")
    quantity = int(input("Quantity : "))
    location = input("Location : ")

    cursor.execute("""
        INSERT INTO inventory(product_name,quantity,location)
        VALUES(?,?,?)
    """, (product, quantity, location))

    conn.commit()

    print("Inventory Added Successfully")
    
def add_sale():

    customer = input("Customer Name : ")
    product = input("Product Name : ")
    quantity = int(input("Quantity : "))

    cursor.execute(
        "SELECT id FROM customers WHERE name=?",
        (customer,)
    )

    if cursor.fetchone() is None:
        print("Customer Not Found")
        return

    cursor.execute(
        "SELECT price, quantity FROM products WHERE name=?",
        (product,)
    )

    data = cursor.fetchone()

    if data is None:
        print("Product Not Found")
        return

    price = data[0]
    stock = data[1]

    if quantity > stock:
        print("Not Enough Stock")
        return

    total = price * quantity

    cursor.execute("""
    INSERT INTO sales(customer,product,quantity,price,total)
    VALUES(?,?,?,?,?)
    """,
    (customer, product, quantity, price, total))

    cursor.execute("""
    UPDATE products
    SET quantity=quantity-?
    WHERE name=?
    """,
    (quantity, product))

    conn.commit()

    print("Sale Completed Successfully")

def view_sales():

    cursor.execute("SELECT * FROM sales")

    rows = cursor.fetchall()

    print("\n========== SALES ==========")

    grand = 0

    for row in rows:

        print("Bill ID :", row[0])
        print("Customer :", row[1])
        print("Product :", row[2])
        print("Quantity :", row[3])
        print("Price :", row[4])
        print("Total :", row[5])

        grand += row[5]

        print("-"*40)

    print("Grand Total Sales =", grand)

def generate_bill():

    customer = input("Customer Name : ")

    cursor.execute(
        "SELECT * FROM sales WHERE customer=?",
        (customer,)
    )

    rows = cursor.fetchall()

    if len(rows) == 0:
        print("No Bill Found")
        return

    subtotal = 0

    print("\n========== INVOICE ==========")
    print("Customer :", customer)
    print("-" * 40)

    for row in rows:

        print(row[2], "x", row[3], "=", row[5])

        subtotal += row[5]

    gst = subtotal * 0.18

    grand = subtotal + gst

    print("-" * 40)
    print("Subtotal :", subtotal)
    print("GST 18% :", gst)
    print("Grand Total :", grand)


def view_inventory():

    cursor.execute("SELECT * FROM inventory")

    rows = cursor.fetchall()

    print("\n====== INVENTORY ======")

    for row in rows:

        print("ID :", row[0])
        print("Product :", row[1])
        print("Quantity :", row[2])
        print("Location :", row[3])
        print("-" * 40)
        
def update_inventory():

    inv_id = int(input("Inventory ID : "))

    product = input("New Product Name : ")
    quantity = int(input("New Quantity : "))
    location = input("New Location : ")

    cursor.execute("""
    UPDATE inventory
    SET product_name=?, quantity=?, location=?
    WHERE id=?
    """, (product, quantity, location, inv_id))

    conn.commit()

    print("Inventory Updated Successfully")


def delete_inventory():

    inv_id = int(input("Inventory ID : "))

    cursor.execute(
        "DELETE FROM inventory WHERE id=?",
        (inv_id,)
    )

    conn.commit()

    print("Inventory Deleted Successfully")

def search_inventory():

    product = input("Enter Product Name : ")

    cursor.execute(
        "SELECT * FROM inventory WHERE product_name LIKE ?",
        ("%" + product + "%",)
    )

    rows = cursor.fetchall()

    if len(rows) == 0:
        print("No Product Found")
        return

    print("\n======= SEARCH RESULT =======")

    for row in rows:

        print("ID :", row[0])
        print("Product :", row[1])
        print("Quantity :", row[2])
        print("Location :", row[3])
        print("-" * 40)
import datetime

def add_payment():

    customer = input("Customer Name : ")

    cursor.execute(
        "SELECT * FROM customers WHERE name=?",
        (customer,)
    )

    if cursor.fetchone() is None:
        print("Customer Not Found")
        return

    amount = float(input("Amount : "))

    print("\n1. Cash")
    print("2. UPI")
    print("3. Card")

    choice = input("Choose Payment Method : ")

    if choice == "1":
        method = "Cash"
    elif choice == "2":
        method = "UPI"
    elif choice == "3":
        method = "Card"
    else:
        print("Invalid Method")
        return

    date = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    cursor.execute("""
    INSERT INTO payments(customer,amount,payment_method,payment_date)
    VALUES(?,?,?,?)
    """,(customer,amount,method,date))

    conn.commit()

    print("Payment Added Successfully")


def view_payments():

    cursor.execute("SELECT * FROM payments")

    rows = cursor.fetchall()

    if len(rows) == 0:
        print("No Payments Found")
        return

    total = 0

    print("\n========== PAYMENT HISTORY ==========")

    for row in rows:

        print("Payment ID :", row[0])
        print("Customer :", row[1])
        print("Amount :", row[2])
        print("Method :", row[3])
        print("Date :", row[4])

        total += row[2]

        print("-" * 40)

    print("Total Payment Received =", total)

def search_payment():

    customer = input("Customer Name : ")

    cursor.execute(
        "SELECT * FROM payments WHERE customer LIKE ?",
        ("%"+customer+"%",)
    )

    rows = cursor.fetchall()

    if len(rows)==0:
        print("No Payment Found")
        return

    for row in rows:
        print(row)

def delete_payment():

    pid = int(input("Payment ID : "))

    cursor.execute(
        "DELETE FROM payments WHERE id=?",
        (pid,)
    )

    conn.commit()

    print("Payment Deleted")


def payment_report():

    cursor.execute("SELECT * FROM payments")

    rows = cursor.fetchall()

    if len(rows) == 0:
        print("No Payment Records Found")
        return

    total_amount = 0
    cash_total = 0
    upi_total = 0
    card_total = 0

    print("\n========== PAYMENT REPORT ==========")
    print(f"{'ID':<5} {'Customer':<15} {'Amount':<10} {'Method':<10} {'Date':<20}")
    print("-" * 70)

    for row in rows:
        print(f"{row[0]:<5} {row[1]:<15} {row[2]:<10} {row[3]:<10} {row[4]:<20}")

        total_amount += row[2]

        if row[3] == "Cash":
            cash_total += row[2]
        elif row[3] == "UPI":
            upi_total += row[2]
        elif row[3] == "Card":
            card_total += row[2]

    print("-" * 70)
    print(f"Total Payment : {total_amount}")
    print(f"Cash : {cash_total}")
    print(f"UPI : {upi_total}")
    print(f"Card : {card_total}")

import datetime

def mark_attendance():

    employee = input("Employee Name : ")

    cursor.execute(
        "SELECT * FROM employees WHERE name=?",
        (employee,)
    )

    if cursor.fetchone() is None:
        print("Employee Not Found")
        return

    print("1. Present")
    print("2. Absent")
    print("3. Leave")

    ch = input("Status : ")

    if ch == "1":
        status = "Present"
    elif ch == "2":
        status = "Absent"
    elif ch == "3":
        status = "Leave"
    else:
        print("Invalid Status")
        return

    today = datetime.date.today()

    cursor.execute("""
    INSERT INTO attendance(employee_name,date,status)
    VALUES(?,?,?)
    """,
    (employee,str(today),status))

    conn.commit()

    print("Attendance Marked Successfully")

def view_attendance():

    cursor.execute("SELECT * FROM attendance")

    rows = cursor.fetchall()

    print("\n========== ATTENDANCE ==========")

    for row in rows:

        print("ID :",row[0])
        print("Employee :",row[1])
        print("Date :",row[2])
        print("Status :",row[3])
        print("-"*40)
    
def search_attendance():

    name = input("Employee Name : ")

    cursor.execute(
        "SELECT * FROM attendance WHERE employee_name LIKE ?",
        ("%"+name+"%",)
    )

    rows = cursor.fetchall()

    if len(rows)==0:
        print("Record Not Found")
        return

    for row in rows:
        print(row)

def attendance_report():

    cursor.execute("""
    SELECT status,COUNT(*)
    FROM attendance
    GROUP BY status
    """)

    rows = cursor.fetchall()

    print("\nAttendance Report")

    for row in rows:

        print(row[0],":",row[1])

def add_payroll():

    employee = input("Employee Name : ")

    cursor.execute(
        "SELECT * FROM employees WHERE name=?",
        (employee,)
    )

    if cursor.fetchone() is None:
        print("Employee Not Found")
        return

    basic = float(input("Basic Salary : "))
    bonus = float(input("Bonus : "))
    deduction = float(input("Deduction : "))

    net = basic + bonus - deduction

    cursor.execute("""
    INSERT INTO payroll(
        employee_name,
        basic_salary,
        bonus,
        deduction,
        net_salary
    )
    VALUES(?,?,?,?,?)
    """,
    (employee,basic,bonus,deduction,net))

    conn.commit()

    print("Payroll Added Successfully")

def view_payroll():

    cursor.execute("SELECT * FROM payroll")

    rows = cursor.fetchall()

    print("\n========== PAYROLL ==========")

    for row in rows:

        print("Payroll ID :",row[0])
        print("Employee :",row[1])
        print("Basic Salary :",row[2])
        print("Bonus :",row[3])
        print("Deduction :",row[4])
        print("Net Salary :",row[5])

        print("-"*40)
        
def search_payroll():

    name = input("Employee Name : ")

    cursor.execute(
        "SELECT * FROM payroll WHERE employee_name LIKE ?",
        ("%"+name+"%",)
    )

    rows = cursor.fetchall()

    if len(rows)==0:
        print("Payroll Not Found")
        return

    for row in rows:
        print(row)

def delete_payroll():

    pid = int(input("Payroll ID : "))

    cursor.execute(
        "DELETE FROM payroll WHERE id=?",
        (pid,)
    )

    conn.commit()

    print("Payroll Deleted")


def payroll_report():

    cursor.execute("""
    SELECT employee_name,net_salary
    FROM payroll
    """)

    rows = cursor.fetchall()

    total = 0

    print("\n========== PAYROLL REPORT ==========")

    for row in rows:

        print("Employee :",row[0])
        print("Net Salary :",row[1])

        total += row[1]

        print("-"*40)

    print("Total Salary Paid =",total)

def dashboard():

    cursor.execute("SELECT COUNT(*) FROM employees")
    employees = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM customers")
    customers = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM suppliers")
    suppliers = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM products")
    products = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM inventory")
    inventory = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM sales")
    sales = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM payments")
    payments = cursor.fetchone()[0]

    cursor.execute("SELECT IFNULL(SUM(total),0) FROM sales")
    sales_amount = cursor.fetchone()[0]

    cursor.execute("SELECT IFNULL(SUM(amount),0) FROM payments")
    payment_amount = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM attendance")
    attendance = cursor.fetchone()[0]

    cursor.execute("SELECT IFNULL(SUM(net_salary),0) FROM payroll")
    payroll = cursor.fetchone()[0]

    print("\n")
    print("="*50)
    print("          ERP DASHBOARD")
    print("="*50)

    print("Employees           :", employees)
    print("Customers           :", customers)
    print("Suppliers           :", suppliers)
    print("Products            :", products)
    print("Inventory Items     :", inventory)
    print("Sales Records       :", sales)
    print("Payments            :", payments)
    print("Attendance Records  :", attendance)

    print("-"*50)

    print("Total Sales Amount  :", sales_amount)
    print("Total Payments      :", payment_amount)
    print("Payroll Paid        :", payroll)

    print("="*50)

def export_employees_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Employees"

    ws.append([
        "Employee ID",
        "Name",
        "Department",
        "Salary"
    ])

    cursor.execute("SELECT * FROM employees")

    rows = cursor.fetchall()

    for row in rows:
        ws.append(row)

    wb.save("Employees_Report.xlsx")

    print("Employees exported successfully.")

def export_customers_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Customers"

    ws.append([
        "Customer ID",
        "Name",
        "Phone"
    ])

    cursor.execute("SELECT * FROM customers")

    rows = cursor.fetchall()

    for row in rows:
        ws.append(row)

    wb.save("Customers_Report.xlsx")

    print("Customers exported successfully.")

def export_products_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Products"

    ws.append([
        "Product ID",
        "Name",
        "Price",
        "Stock"
    ])

    cursor.execute("SELECT * FROM products")

    rows = cursor.fetchall()

    for row in rows:
        ws.append(row)

    wb.save("Products_Report.xlsx")

    print("Products exported successfully.")

def export_sales_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Sales"

    ws.append([
        "Bill ID",
        "Customer",
        "Product",
        "Quantity",
        "Price",
        "Total"
    ])

    cursor.execute("SELECT * FROM sales")

    rows = cursor.fetchall()

    for row in rows:
        ws.append(row)

    wb.save("Sales_Report.xlsx")

    print("Sales exported successfully.")

def export_inventory_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventory"

    ws.append([
        "Inventory ID",
        "Product",
        "Quantity",
        "Location"
    ])

    cursor.execute("SELECT * FROM inventory")

    rows = cursor.fetchall()

    for row in rows:
        ws.append(row)

    wb.save("Inventory_Report.xlsx")

    print("Inventory exported successfully.")

def dashboard_chart():

    labels = [
        "Employees",
        "Customers",
        "Products",
        "Suppliers",
        "Sales",
        "Payments"
    ]

    cursor.execute("SELECT COUNT(*) FROM employees")
    employees = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM customers")
    customers = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM products")
    products = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM suppliers")
    suppliers = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM sales")
    sales = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM payments")
    payments = cursor.fetchone()[0]

    values = [
        employees,
        customers,
        products,
        suppliers,
        sales,
        payments
    ]

    plt.figure(figsize=(10,6))
    plt.bar(labels, values)

    plt.title("ERP Dashboard")

    plt.xlabel("Modules")

    plt.ylabel("Records")

    plt.show()

def sales_pie_chart():

    cursor.execute("""
    SELECT product,
           SUM(quantity)
    FROM sales
    GROUP BY product
    """)

    rows = cursor.fetchall()

    if len(rows) == 0:
        print("No Sales Found")
        return

    labels = []
    values = []

    for row in rows:
        labels.append(row[0])
        values.append(row[1])

    plt.figure(figsize=(7,7))
    plt.pie(values,
            labels=labels,
            autopct="%1.1f%%")

    plt.title("Product Sales")

    plt.show()

def payment_chart():

    cursor.execute("""
    SELECT payment_method,
           SUM(amount)
    FROM payments
    GROUP BY payment_method
    """)

    rows = cursor.fetchall()

    labels = []
    values = []

    for row in rows:
        labels.append(row[0])
        values.append(row[1])

    plt.figure(figsize=(7,5))

    plt.bar(labels, values)

    plt.title("Payment Methods")

    plt.xlabel("Method")

    plt.ylabel("Amount")

    plt.show()

def dashboard():

    cursor.execute("SELECT COUNT(*) FROM employees")
    emp = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM customers")
    cus = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM suppliers")
    sup = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM products")
    pro = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM sales")
    sale = cursor.fetchone()[0]

    cursor.execute("SELECT IFNULL(SUM(total),0) FROM sales")
    revenue = cursor.fetchone()[0]

    cursor.execute("SELECT IFNULL(SUM(amount),0) FROM payments")
    payment = cursor.fetchone()[0]

    print("="*50)
    print("        ERP DASHBOARD")
    print("="*50)

    print("Employees        :", emp)
    print("Customers        :", cus)
    print("Suppliers        :", sup)
    print("Products         :", pro)
    print("Sales            :", sale)
    print("Revenue          :", revenue)
    print("Payments         :", payment)

    print("="*50)

# ---------------- MENU ----------------

while True:

    print("\n========== ERP MANAGEMENT SYSTEM ==========")

    # Employee
    print("1  Add Employee")
    print("2  View Employee")
    print("3  Update Employee")
    print("4  Delete Employee")

    # Customer
    print("5  Add Customer")
    print("6  View Customer")

    # Product
    print("7  Add Product")
    print("8  View Product")
    print("9  Search Product")
    print("10 Update Product")
    print("11 Delete Product")

    # Inventory
    print("12 Add Inventory")
    print("13 View Inventory")
    print("14 Update Inventory")
    print("15 Delete Inventory")
    print("16 Search Inventory")

    # Sales
    print("17 Add Sale")
    print("18 View Sales")
    print("19 Generate Bill")

    # Supplier
    print("20 Add Supplier")
    print("21 View Supplier")
    print("22 Update Supplier")
    print("23 Delete Supplier")
    print("24 Search Supplier")

    # Payments
    print("25 Add Payment")
    print("26 View Payments")
    print("27 Search Payment")
    print("28 Delete Payment")
    print("29 Payment Report")

    # Attendance
    print("30 Mark Attendance")
    print("31 View Attendance")
    print("32 Search Attendance")
    print("33 Attendance Report")
    
    # Payroll
    print("35 Add Payroll")
    print("36 View Payroll")    
    print("37 Search Payroll")
    print("38 Update Payroll")
    print("39 Delete Payroll")
    print("40 Payroll Report")

    # Dashboard
    print("41 Dashboard")

    # Export to Excel
    print("42 Export Employees Excel")
    print("43 Export Customers Excel")
    print("44 Export Products Excel")
    print("45 Export Inventory Excel")
    print("46 Export Sales Excel")
    print("47 Export Payments Excel")
    print("48 Export Inventory Excel")

    # Charts
    print("49 Dashboard")
    print("50 Dashboard Chart")
    print("51 Sales Pie Chart")
    print("52 Payment Chart")

    # Exit
    print("53 Exit")

    choice = input("\nEnter Choice : ")

    # Employee
    if choice == "1":
        add_employee()

    elif choice == "2":
        view_employee()

    elif choice == "3":
        update_employee()

    elif choice == "4":
        delete_employee()

    # Customer
    elif choice == "5":
        add_customer()

    elif choice == "6":
        view_customer()

    # Product
    elif choice == "7":
        add_product()

    elif choice == "8":
        view_product()

    elif choice == "9":
        search_product()

    elif choice == "10":
        update_product()

    elif choice == "11":
        delete_product()

    # Inventory
    elif choice == "12":
        add_inventory()

    elif choice == "13":
        view_inventory()

    elif choice == "14":
        update_inventory()

    elif choice == "15":
        delete_inventory()

    elif choice == "16":
        search_inventory()

    # Sales
    elif choice == "17":
        add_sale()

    elif choice == "18":
        view_sales()

    elif choice == "19":
        generate_bill()

    # Supplier
    elif choice == "20":
        add_supplier()

    elif choice == "21":
        view_supplier()

    elif choice == "22":
        update_supplier()

    elif choice == "23":
        delete_supplier()

    elif choice == "24":
        search_supplier()

    # Payments
    elif choice == "25":
        add_payment()

    elif choice == "26":
        view_payments()

    elif choice == "27":
        search_payment()

    elif choice == "28":
        delete_payment()

    elif choice == "29":
        payment_report()

    # Attendance
    elif choice == "30":
        mark_attendance()

    elif choice == "31":
        view_attendance()

    elif choice == "32":
        search_attendance()

    elif choice == "33":
        attendance_report()

    # Payroll
    elif choice == "35":    
        add_payroll()

    elif choice == "36":
        view_payroll() 

    elif choice == "37":
        search_payroll()

    elif choice == "38":
        print("Payroll Update Functionality Not Implemented Yet") 

    elif choice == "39":
        delete_payroll()    
    
    elif choice == "40":
        payroll_report()

    # Dashboard
    elif choice == "41":
        dashboard() 
    
    # Export to Excel
    elif choice == "42":
        export_employees_excel()

    elif choice == "43":
        export_customers_excel()

    elif choice == "44":
        export_products_excel()

    elif choice == "45":
        export_inventory_excel()

    elif choice == "46":
        export_sales_excel()

    elif choice == "47":
        print("Export Payments Excel Functionality Not Implemented Yet")

    elif choice == "48":
        print("Export Inventory Excel Functionality Not Implemented Yet")
    
    # Charts
    elif choice == "49":
        dashboard()
    elif choice == "50":
        dashboard_chart()
    elif choice == "51":
        sales_pie_chart()
    elif choice == "52":
        payment_chart()

    # Exit
    elif choice == "53":
        conn.close()
        print("Thank You for using ERP Management System.")
        break

    else:
        print("Invalid Choice")
       